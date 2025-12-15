from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.styles import Font, Alignment


def _write_table(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> tuple[int, int]:
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

    last_row = start_row + len(df)  # header + data
    last_col = start_col + len(df.columns) - 1

    # Header styling
    header_font = Font(bold=True)
    for c in range(start_col, last_col + 1):
        cell = ws.cell(row=start_row, column=c)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    return last_row, last_col


def build_excel_report(
    *,
    output_path: Path,
    kpis: dict[str, float],
    df_daily: pd.DataFrame,
    df_channel: pd.DataFrame,
    df_products: pd.DataFrame,
) -> Path:
    wb = Workbook()

    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum["A1"] = "KPI"
    ws_sum["B1"] = "Value"
    ws_sum["A1"].font = Font(bold=True)
    ws_sum["B1"].font = Font(bold=True)

    kpi_rows = [
        ("Total Orders", kpis["total_orders"]),
        ("Total Customers", kpis["total_customers"]),
        ("Total Net Revenue", kpis["total_net_revenue"]),
        ("Average Order Value", kpis["avg_order_value"]),
    ]
    for i, (k, v) in enumerate(kpi_rows, start=2):
        ws_sum[f"A{i}"] = k
        ws_sum[f"B{i}"] = float(v)

    ws_sum.column_dimensions["A"].width = 25
    ws_sum.column_dimensions["B"].width = 18

    # Daily sheet
    ws_daily = wb.create_sheet("Revenue by Day")
    last_row, _ = _write_table(ws_daily, df_daily)

    chart = LineChart()
    chart.title = "Net Revenue by Day"
    chart.y_axis.title = "Net Revenue"
    chart.x_axis.title = "Day"

    data = Reference(ws_daily, min_col=2, min_row=1, max_row=last_row)
    cats = Reference(ws_daily, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws_daily.add_chart(chart, "D2")

    # Channel sheet
    ws_channel = wb.create_sheet("Revenue by Channel")
    last_row_c, _ = _write_table(ws_channel, df_channel)

    bar = BarChart()
    bar.title = "Net Revenue by Channel"
    bar.y_axis.title = "Net Revenue"
    bar.x_axis.title = "Channel"

    data_c = Reference(ws_channel, min_col=2, min_row=1, max_row=last_row_c)
    cats_c = Reference(ws_channel, min_col=1, min_row=2, max_row=last_row_c)
    bar.add_data(data_c, titles_from_data=True)
    bar.set_categories(cats_c)
    ws_channel.add_chart(bar, "D2")

    # Products sheet
    ws_prod = wb.create_sheet("Top Products")
    _write_table(ws_prod, df_products)

    wb.save(output_path)
    return output_path
